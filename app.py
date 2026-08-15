from __future__ import annotations

import io
import json
import os
import re
import sqlite3
import hashlib
import secrets
import requests
from contextlib import contextmanager
from datetime import datetime, timezone
from pathlib import Path
from uuid import uuid4

from authlib.integrations.starlette_client import OAuth
from cryptography.fernet import Fernet
from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import FileResponse, RedirectResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request as GoogleRequest
from googleapiclient.discovery import build
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field
from starlette.middleware.sessions import SessionMiddleware

ROOT = Path(__file__).parent
DB = Path(os.getenv("DATABASE_PATH", ROOT / "everafter.sqlite3"))
STATIC = ROOT / "static"
app = FastAPI(title="Everafter")
app.add_middleware(SessionMiddleware, secret_key=os.getenv("SESSION_SECRET", "development-session-secret-change-me"), https_only=os.getenv("COOKIE_SECURE", "false").lower() == "true", same_site="lax")
app.mount("/static", StaticFiles(directory=STATIC), name="static")
GOOGLE_SCOPES = "openid email profile https://www.googleapis.com/auth/drive.readonly"
GOOGLE_CLIENT_ID = os.getenv("GOOGLE_CLIENT_ID")
GOOGLE_CLIENT_SECRET = os.getenv("GOOGLE_CLIENT_SECRET")
GOOGLE_REDIRECT_URI = os.getenv("GOOGLE_REDIRECT_URI")
GOOGLE_PICKER_API_KEY = os.getenv("GOOGLE_PICKER_API_KEY")
TOKEN_KEY = os.getenv("TOKEN_ENCRYPTION_KEY")
oauth = OAuth()
if GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET:
    oauth.register("google", client_id=GOOGLE_CLIENT_ID, client_secret=GOOGLE_CLIENT_SECRET, server_metadata_url="https://accounts.google.com/.well-known/openid-configuration", client_kwargs={"scope": GOOGLE_SCOPES})

@contextmanager
def db():
    connection = sqlite3.connect(DB)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    try:
        yield connection
        connection.commit()
    finally:
        connection.close()

def now() -> str:
    return datetime.now(timezone.utc).isoformat()

def one(connection, query: str, values=()):
    return connection.execute(query, values).fetchone()

def require_project(connection, project_id: str, user_id: str = "user_emily"):
    row = one(connection, """SELECT p.*, pm.role FROM projects p
        JOIN project_members pm ON pm.project_id=p.id
        WHERE p.id=? AND pm.user_id=?""", (project_id, user_id))
    if not row:
        raise HTTPException(404, "Project not found")
    return row

def current_user_id(request: Request) -> str:
    user_id = request.session.get("user_id")
    if user_id:
        return user_id
    raise HTTPException(401, "Sign in with Google to continue")

def cipher() -> Fernet:
    if not TOKEN_KEY:
        raise HTTPException(503, "Server token encryption is not configured")
    return Fernet(TOKEN_KEY.encode())

def high_quality_thumbnail(url: str | None) -> str | None:
    """Google Drive thumbnail URLs normally end in =s220; 600px is crisp in the gallery."""
    if not url:
        return url
    return re.sub(r"=s\d+(?:-[A-Za-z0-9-]+)?$", "=s600", url)

def drive_client_for_user(user_id: str):
    """Refreshes the owner's Drive access token and returns an authenticated client."""
    if not GOOGLE_CLIENT_ID or not GOOGLE_CLIENT_SECRET:
        raise HTTPException(503, "Google OAuth is not configured")
    with db() as c:
        stored = one(c, "SELECT google_token_encrypted FROM users WHERE id=?", (user_id,))
        if not stored or not stored["google_token_encrypted"]:
            raise HTTPException(401, "Reconnect your Google account")
        token = json.loads(cipher().decrypt(stored["google_token_encrypted"].encode()).decode())
        credentials = Credentials(token=token.get("access_token"), refresh_token=token.get("refresh_token"), token_uri="https://oauth2.googleapis.com/token", client_id=GOOGLE_CLIENT_ID, client_secret=GOOGLE_CLIENT_SECRET, scopes=GOOGLE_SCOPES.split())
        if credentials.refresh_token:
            credentials.refresh(GoogleRequest())
            token["access_token"] = credentials.token
            c.execute("UPDATE users SET google_token_encrypted=? WHERE id=?", (cipher().encrypt(json.dumps(token).encode()).decode(),user_id))
    return build("drive", "v3", credentials=credentials, cache_discovery=False), credentials.token

def init_db():
    with db() as c:
        c.executescript("""
        CREATE TABLE IF NOT EXISTS users (id TEXT PRIMARY KEY, google_sub TEXT UNIQUE, email TEXT UNIQUE NOT NULL, display_name TEXT NOT NULL, avatar_initials TEXT NOT NULL, google_token_encrypted TEXT);
        CREATE TABLE IF NOT EXISTS projects (id TEXT PRIMARY KEY, owner_id TEXT NOT NULL REFERENCES users(id), name TEXT NOT NULL, drive_folder_id TEXT, general_note TEXT NOT NULL DEFAULT '', created_at TEXT NOT NULL, updated_at TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS project_members (project_id TEXT REFERENCES projects(id) ON DELETE CASCADE, user_id TEXT REFERENCES users(id) ON DELETE CASCADE, role TEXT NOT NULL CHECK(role IN ('owner','collaborator')), PRIMARY KEY(project_id,user_id));
        CREATE TABLE IF NOT EXISTS photos (id TEXT PRIMARY KEY, project_id TEXT NOT NULL REFERENCES projects(id) ON DELETE CASCADE, drive_file_id TEXT NOT NULL, file_name TEXT NOT NULL, web_view_link TEXT, thumbnail_url TEXT NOT NULL, is_available INTEGER NOT NULL DEFAULT 1, UNIQUE(project_id, drive_file_id), UNIQUE(project_id,file_name));
        CREATE TABLE IF NOT EXISTS albums (id TEXT PRIMARY KEY, project_id TEXT NOT NULL REFERENCES projects(id) ON DELETE CASCADE, parent_album_id TEXT REFERENCES albums(id), name TEXT NOT NULL, is_source INTEGER NOT NULL DEFAULT 0, created_by TEXT NOT NULL REFERENCES users(id), created_at TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS album_photos (album_id TEXT REFERENCES albums(id) ON DELETE CASCADE, photo_id TEXT REFERENCES photos(id) ON DELETE CASCADE, liked INTEGER NOT NULL DEFAULT 0, updated_at TEXT NOT NULL, PRIMARY KEY(album_id,photo_id));
        CREATE TABLE IF NOT EXISTS photo_notes (id TEXT PRIMARY KEY, photo_id TEXT NOT NULL REFERENCES photos(id) ON DELETE CASCADE, content TEXT NOT NULL, created_by TEXT NOT NULL REFERENCES users(id), created_at TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS project_invites (id TEXT PRIMARY KEY, project_id TEXT NOT NULL REFERENCES projects(id) ON DELETE CASCADE, token_hash TEXT NOT NULL UNIQUE, created_by TEXT NOT NULL REFERENCES users(id), created_at TEXT NOT NULL);
        """)
        columns = {row["name"] for row in c.execute("PRAGMA table_info(users)")}
        if "google_token_encrypted" not in columns:
            c.execute("ALTER TABLE users ADD COLUMN google_token_encrypted TEXT")
        if one(c, "SELECT id FROM users LIMIT 1") or os.getenv("DEMO_MODE", "false").lower() != "true":
            return
        users = [("user_emily", "google-emily-demo", "emily@example.com", "Emily Tran", "ET"), ("user_james", "google-james-demo", "james@example.com", "James Nguyen", "JN")]
        c.executemany("INSERT INTO users(id,google_sub,email,display_name,avatar_initials) VALUES (?,?,?,?,?)", users)
        project_id, source_id, shortlist_id = "project_demo", "album_source", "album_shortlist"
        c.execute("INSERT INTO projects VALUES (?,?,?,?,?,?,?)", (project_id,"user_emily","Emily & James — Pre-wedding","drive-folder-demo","Keep the edits natural and warm. Please preserve the golden-hour tones across the selected photos.",now(),now()))
        c.executemany("INSERT INTO project_members VALUES (?,?,?)", [(project_id,"user_emily","owner"),(project_id,"user_james","collaborator")])
        c.executemany("INSERT INTO albums VALUES (?,?,?,?,?,?,?)", [(source_id,project_id,None,"All photos",1,"user_emily",now()),(shortlist_id,project_id,source_id,"Shortlist round 2",0,"user_emily",now())])
        pictures = [
          ("DSC_0284.jpg","https://images.unsplash.com/photo-1519741497674-611481863552?auto=format&fit=crop&w=1000&q=85"),
          ("DSC_0285.jpg","https://images.unsplash.com/photo-1606800052052-a08af7148866?auto=format&fit=crop&w=1000&q=85"),
          ("DSC_0291.jpg","https://images.unsplash.com/photo-1511285560929-80b456fea0bc?auto=format&fit=crop&w=1000&q=85"),
          ("DSC_0308.jpg","https://images.unsplash.com/photo-1519225421980-715cb0215aed?auto=format&fit=crop&w=1000&q=85"),
          ("DSC_0314.jpg","https://images.unsplash.com/photo-1464366400600-7168b8af9bc3?auto=format&fit=crop&w=1000&q=85"),
          ("DSC_0332.jpg","https://images.unsplash.com/photo-1522673607200-164d1b6ce486?auto=format&fit=crop&w=1000&q=85"),
          ("DSC_0348.jpg","https://images.unsplash.com/photo-1544078751-58fee2d8a03b?auto=format&fit=crop&w=1000&q=85"),
          ("DSC_0359.jpg","https://images.unsplash.com/photo-1515934751635-c81c6bc9a2d8?auto=format&fit=crop&w=1000&q=85"),
        ]
        for index, (name, url) in enumerate(pictures):
            photo_id = f"photo_{index+1}"
            c.execute("INSERT INTO photos VALUES (?,?,?,?,?,?,?)", (photo_id,project_id,f"drive_{index+1}",name,"https://drive.google.com/file/d/drive_{index+1}/view",url,1))
            c.execute("INSERT INTO album_photos VALUES (?,?,?,?)", (source_id,photo_id,1 if index < 5 else 0,now()))
            c.execute("INSERT INTO album_photos VALUES (?,?,?,?)", (shortlist_id,photo_id,1 if index in (0,1,3,5) else 0,now()))
        c.execute("INSERT INTO photo_notes VALUES (?,?,?,?,?)", ("note_1","photo_1","Keep the skin texture natural; soften the shadow on the left.","user_emily",now()))

class ProjectInput(BaseModel): name: str = Field(min_length=1, max_length=100)
class AlbumInput(BaseModel): name: str = Field(min_length=1, max_length=100); parent_album_id: str
class NoteInput(BaseModel): content: str = Field(min_length=1, max_length=1000)
class GeneralNoteInput(BaseModel): content: str = Field(max_length=3000)
class DriveFolderInput(BaseModel): folder_url_or_id: str = Field(min_length=10, max_length=1000)

@app.get("/")
def home(): return FileResponse(STATIC / "index.html")

@app.get("/healthz")
def healthz(): return {"status": "ok"}

@app.get("/auth/google/login")
async def google_login(request: Request):
    if not GOOGLE_CLIENT_ID or not GOOGLE_CLIENT_SECRET or not GOOGLE_REDIRECT_URI:
        raise HTTPException(503, "Google OAuth is not configured on this server")
    return await oauth.google.authorize_redirect(request, GOOGLE_REDIRECT_URI, access_type="offline", prompt="consent")

@app.get("/auth/google/callback")
async def google_callback(request: Request):
    if not TOKEN_KEY:
        raise HTTPException(503, "TOKEN_ENCRYPTION_KEY is not configured")
    token = await oauth.google.authorize_access_token(request)
    profile = token.get("userinfo")
    if not profile:
        profile_response = await oauth.google.get("https://openidconnect.googleapis.com/v1/userinfo", token=token)
        profile = profile_response.json()
    google_sub, email = profile["sub"], profile["email"]
    user_id = f"google_{google_sub}"
    initials = "".join(part[0] for part in profile.get("name", email).split()[:2]).upper()
    encrypted = cipher().encrypt(json.dumps(token).encode()).decode()
    with db() as c:
        c.execute("""INSERT INTO users(id,google_sub,email,display_name,avatar_initials,google_token_encrypted)
          VALUES (?,?,?,?,?,?) ON CONFLICT(google_sub) DO UPDATE SET email=excluded.email,display_name=excluded.display_name,avatar_initials=excluded.avatar_initials,google_token_encrypted=excluded.google_token_encrypted""", (user_id,google_sub,email,profile.get("name",email),initials,encrypted))
    request.session["user_id"] = user_id
    invite_token = request.session.pop("pending_invite", None)
    if invite_token:
        accept_invite_token(invite_token, user_id)
    return RedirectResponse("/")

@app.post("/auth/logout")
def logout(request: Request):
    request.session.clear()
    return {"ok": True}

@app.post("/auth/demo-login")
def demo_login(request: Request):
    if os.getenv("DEMO_MODE", "false").lower() != "true":
        raise HTTPException(404, "Demo sign-in is disabled")
    request.session["user_id"] = "user_emily"
    invite_token = request.session.pop("pending_invite", None)
    if invite_token:
        accept_invite_token(invite_token, "user_emily")
    return {"ok": True}

def invite_hash(token: str) -> str:
    return hashlib.sha256(token.encode()).hexdigest()

def accept_invite_token(token: str, user_id: str) -> bool:
    with db() as c:
        invite = one(c, "SELECT project_id FROM project_invites WHERE token_hash=?", (invite_hash(token),))
        if not invite:
            return False
        c.execute("INSERT OR IGNORE INTO project_members(project_id,user_id,role) VALUES (?,?,?)", (invite["project_id"],user_id,"collaborator"))
    return True

@app.get("/invite/{token}")
def open_invite(token: str, request: Request):
    request.session["pending_invite"] = token
    user_id = request.session.get("user_id")
    if user_id:
        accept_invite_token(token, user_id)
        request.session.pop("pending_invite", None)
    return RedirectResponse("/")

@app.get("/api/google/picker-config")
def google_picker_config(request: Request):
    user_id = current_user_id(request)
    if not GOOGLE_PICKER_API_KEY:
        raise HTTPException(503, "GOOGLE_PICKER_API_KEY is not configured")
    _, access_token = drive_client_for_user(user_id)
    return {"developer_key": GOOGLE_PICKER_API_KEY, "oauth_token": access_token}

@app.get("/api/me")
def me(request: Request):
    with db() as c:
        user = one(c, "SELECT id,email,display_name,avatar_initials FROM users WHERE id=?", (current_user_id(request),))
        if not user: raise HTTPException(401, "Session user not found")
        return dict(user)

@app.get("/api/projects")
def projects(request: Request):
    user_id = current_user_id(request)
    with db() as c:
        rows = c.execute("""SELECT p.id,p.name,p.owner_id,p.updated_at,pm.role,
          (SELECT COUNT(*) FROM photos WHERE project_id=p.id) photo_count
          FROM projects p JOIN project_members pm ON pm.project_id=p.id WHERE pm.user_id=? ORDER BY p.updated_at DESC""", (user_id,)).fetchall()
        return [dict(row) for row in rows]

@app.post("/api/projects")
def create_project(data: ProjectInput, request: Request):
    user_id = current_user_id(request)
    project_id, album_id = str(uuid4()), str(uuid4())
    with db() as c:
        c.execute("INSERT INTO projects VALUES (?,?,?,?,?,?,?)", (project_id,user_id,data.name,None,"",now(),now()))
        c.execute("INSERT INTO project_members VALUES (?,?,?)", (project_id,user_id,"owner"))
        c.execute("INSERT INTO albums VALUES (?,?,?,?,?,?,?)", (album_id,project_id,None,"All photos",1,user_id,now()))
    return {"id": project_id}

@app.delete("/api/projects/{project_id}")
def delete_project(project_id: str, request: Request):
    user_id = current_user_id(request)
    with db() as c:
        project = require_project(c, project_id, user_id)
        if project["role"] != "owner":
            raise HTTPException(403, "Only the owner can delete a project")
        c.execute("DELETE FROM projects WHERE id=?", (project_id,))
    return {"ok": True}

@app.post("/api/projects/{project_id}/invites")
def create_invite(project_id: str, request: Request):
    user_id = current_user_id(request)
    with db() as c:
        project = require_project(c, project_id, user_id)
        if project["role"] != "owner":
            raise HTTPException(403, "Only the project owner can invite collaborators")
        token = secrets.token_urlsafe(32)
        c.execute("INSERT INTO project_invites VALUES (?,?,?,?,?)", (str(uuid4()),project_id,invite_hash(token),user_id,now()))
    return {"url": str(request.base_url).rstrip("/") + "/invite/" + token}

@app.post("/api/projects/{project_id}/drive-folder")
def connect_drive_folder(project_id: str, data: DriveFolderInput, request: Request):
    """Imports direct image children only; Drive originals always remain untouched."""
    user_id = current_user_id(request)
    folder_id = re.search(r"(?:folders/)?([A-Za-z0-9_-]{10,})", data.folder_url_or_id)
    if not folder_id: raise HTTPException(400, "Enter a Google Drive folder URL or ID")
    folder_id = folder_id.group(1)
    with db() as c:
        project = require_project(c, project_id, user_id)
        if project["role"] != "owner": raise HTTPException(403, "Only the owner can connect Google Drive")
    drive, _ = drive_client_for_user(user_id)
    try:
        drive.files().get(fileId=folder_id, fields="id,name,mimeType").execute()
        images=[]; page_token=None
        while True:
            result = drive.files().list(q=f"'{folder_id}' in parents and trashed = false", fields="nextPageToken,files(id,name,mimeType,thumbnailLink,webViewLink)", pageSize=1000, orderBy="name", pageToken=page_token).execute()
            images.extend(f for f in result.get("files", []) if f.get("mimeType", "").startswith("image/"))
            page_token=result.get("nextPageToken")
            if not page_token: break
    except Exception as error:
        raise HTTPException(400, f"Google Drive could not read this folder: {error}")
    with db() as c:
        c.execute("UPDATE projects SET drive_folder_id=?,updated_at=? WHERE id=?", (folder_id,now(),project_id))
        source = one(c, "SELECT id FROM albums WHERE project_id=? AND is_source=1", (project_id,))
        for image in images:
            photo_id = str(uuid4())
            c.execute("""INSERT OR IGNORE INTO photos(id,project_id,drive_file_id,file_name,web_view_link,thumbnail_url,is_available)
              VALUES (?,?,?,?,?,?,1)""", (photo_id,project_id,image["id"],image["name"],image.get("webViewLink"),high_quality_thumbnail(image.get("thumbnailLink")) or image.get("webViewLink")))
            saved = one(c,"SELECT id FROM photos WHERE project_id=? AND drive_file_id=?",(project_id,image["id"]))
            c.execute("INSERT OR IGNORE INTO album_photos VALUES (?,?,0,?)",(source["id"],saved["id"],now()))
    return {"folder_id":folder_id,"imported":len(images)}

@app.get("/api/projects/{project_id}")
def project_detail(project_id: str, request: Request):
    with db() as c:
        project = dict(require_project(c, project_id, current_user_id(request)))
        project["members"] = [dict(row) for row in c.execute("SELECT u.display_name,u.avatar_initials,pm.role FROM project_members pm JOIN users u ON u.id=pm.user_id WHERE pm.project_id=?", (project_id,))]
        project["albums"] = [dict(row) for row in c.execute("""SELECT a.*,COUNT(ap.photo_id) photo_count,SUM(ap.liked) liked_count FROM albums a LEFT JOIN album_photos ap ON ap.album_id=a.id WHERE a.project_id=? GROUP BY a.id ORDER BY a.is_source DESC,a.created_at""", (project_id,))]
        return project

@app.get("/api/albums/{album_id}/photos")
def album_photos(album_id: str, request: Request, offset: int = 0, limit: int = 10):
    if offset < 0 or limit < 1 or limit > 50:
        raise HTTPException(400, "offset must be positive and limit must be between 1 and 50")
    with db() as c:
        row = one(c, "SELECT p.id FROM projects p JOIN albums a ON a.project_id=p.id JOIN project_members pm ON pm.project_id=p.id WHERE a.id=? AND pm.user_id=?", (album_id,current_user_id(request)))
        if not row: raise HTTPException(404, "Album not found")
        total = one(c, "SELECT COUNT(*) AS count FROM album_photos WHERE album_id=?", (album_id,))["count"]
        rows = [dict(x) for x in c.execute("""SELECT ph.*,ap.liked FROM album_photos ap
          JOIN photos ph ON ph.id=ap.photo_id WHERE ap.album_id=? ORDER BY ph.file_name LIMIT ? OFFSET ?""", (album_id,limit,offset))]
        for row in rows:
            row["thumbnail_url"] = f"/api/photos/{row['id']}/thumbnail"
        return {"photos": rows, "total": total, "has_more": offset + len(rows) < total}

@app.post("/api/projects/{project_id}/albums")
def create_album(project_id: str, data: AlbumInput, request: Request):
    album_id = str(uuid4())
    with db() as c:
        user_id = current_user_id(request); require_project(c, project_id, user_id)
        parent = one(c, "SELECT id FROM albums WHERE id=? AND project_id=?", (data.parent_album_id,project_id))
        if not parent: raise HTTPException(400, "Choose an album from this project")
        c.execute("INSERT INTO albums VALUES (?,?,?,?,?,?,?)", (album_id,project_id,data.parent_album_id,data.name,0,user_id,now()))
        c.execute("""INSERT INTO album_photos(album_id,photo_id,liked,updated_at)
          SELECT ?,photo_id,0,? FROM album_photos WHERE album_id=? AND liked=1""", (album_id,now(),data.parent_album_id))
    return {"id": album_id}

@app.delete("/api/albums/{album_id}")
def delete_album(album_id: str, request: Request):
    with db() as c:
        album = one(c, "SELECT a.*,pm.role FROM albums a JOIN project_members pm ON pm.project_id=a.project_id WHERE a.id=? AND pm.user_id=?", (album_id,current_user_id(request)))
        if not album: raise HTTPException(404, "Album not found")
        if album["is_source"]: raise HTTPException(400, "The source album cannot be deleted")
        c.execute("DELETE FROM albums WHERE id=?", (album_id,))
    return {"ok": True}

@app.post("/api/albums/{album_id}/photos/{photo_id}/toggle")
def toggle_photo(album_id: str, photo_id: str, request: Request):
    with db() as c:
        item = one(c, "SELECT ap.liked FROM album_photos ap JOIN albums a ON a.id=ap.album_id JOIN project_members pm ON pm.project_id=a.project_id WHERE ap.album_id=? AND ap.photo_id=? AND pm.user_id=?", (album_id,photo_id,current_user_id(request)))
        if item is None: raise HTTPException(404, "Photo not found")
        liked = 0 if item["liked"] else 1
        c.execute("UPDATE album_photos SET liked=?,updated_at=? WHERE album_id=? AND photo_id=?", (liked,now(),album_id,photo_id))
    return {"liked": bool(liked)}

@app.post("/api/albums/{album_id}/like-all")
def like_all_photos(album_id: str, request: Request):
    with db() as c:
        album = one(c, """SELECT a.id FROM albums a JOIN project_members pm ON pm.project_id=a.project_id
          WHERE a.id=? AND pm.user_id=?""", (album_id,current_user_id(request)))
        if not album: raise HTTPException(404, "Album not found")
        c.execute("UPDATE album_photos SET liked=1,updated_at=? WHERE album_id=?", (now(),album_id))
        count = one(c,"SELECT COUNT(*) AS count FROM album_photos WHERE album_id=?",(album_id,))["count"]
    return {"liked_count": count}

@app.get("/api/photos/{photo_id}/notes")
def notes(photo_id: str, request: Request):
    with db() as c:
        access = one(c,"SELECT ph.id FROM photos ph JOIN project_members pm ON pm.project_id=ph.project_id WHERE ph.id=? AND pm.user_id=?",(photo_id,current_user_id(request)))
        if not access: raise HTTPException(404,"Photo not found")
        return [dict(row) for row in c.execute("""SELECT n.*,u.display_name,u.avatar_initials FROM photo_notes n JOIN users u ON u.id=n.created_by WHERE n.photo_id=? ORDER BY n.created_at""", (photo_id,))]

@app.get("/api/photos/{photo_id}/image")
def original_photo(photo_id: str, request: Request):
    """Streams the original Drive image only for the full-screen viewer."""
    user_id = current_user_id(request)
    with db() as c:
        photo = one(c, """SELECT ph.drive_file_id,p.owner_id FROM photos ph
          JOIN projects p ON p.id=ph.project_id
          JOIN project_members pm ON pm.project_id=p.id
          WHERE ph.id=? AND pm.user_id=?""", (photo_id,user_id))
        if not photo: raise HTTPException(404, "Photo not found")
    try:
        drive, _ = drive_client_for_user(photo["owner_id"])
        metadata = drive.files().get(fileId=photo["drive_file_id"], fields="mimeType").execute()
        content = drive.files().get_media(fileId=photo["drive_file_id"]).execute()
    except Exception as error:
        raise HTTPException(404, f"Original image is unavailable: {error}")
    return Response(content=content, media_type=metadata.get("mimeType", "image/jpeg"), headers={"Cache-Control": "private, max-age=300"})

@app.get("/api/photos/{photo_id}/thumbnail")
def photo_thumbnail(photo_id: str, request: Request):
    """Serves Drive thumbnails through the owner connection so collaborators need no Drive permission."""
    user_id = current_user_id(request)
    with db() as c:
        photo = one(c, """SELECT ph.thumbnail_url,ph.drive_file_id,p.owner_id FROM photos ph
          JOIN projects p ON p.id=ph.project_id
          JOIN project_members pm ON pm.project_id=p.id
          WHERE ph.id=? AND pm.user_id=?""", (photo_id,user_id))
        if not photo: raise HTTPException(404, "Photo not found")
    try:
        drive, access_token = drive_client_for_user(photo["owner_id"])
        if photo["thumbnail_url"].startswith("http"):
            try:
                result = requests.get(
                    photo["thumbnail_url"],
                    headers={"Authorization": f"Bearer {access_token}"},
                    timeout=30,
                )
                result.raise_for_status()
                return Response(
                    content=result.content,
                    media_type=result.headers.get("Content-Type", "image/jpeg"),
                    headers={"Cache-Control": "private, max-age=300"},
                )
            except requests.RequestException:
                # A Drive thumbnail URL can be session-specific. Use the file stream
                # as a reliable fallback for collaborators without Drive access.
                pass
        content = drive.files().get_media(fileId=photo["drive_file_id"]).execute()
        return Response(content=content, media_type="image/jpeg", headers={"Cache-Control": "private, max-age=300"})
    except Exception as error:
        raise HTTPException(404, f"Thumbnail is unavailable: {error}")

@app.post("/api/photos/{photo_id}/notes")
def add_note(photo_id: str, data: NoteInput, request: Request):
    note_id = str(uuid4())
    with db() as c:
        user_id=current_user_id(request)
        if not one(c,"SELECT ph.id FROM photos ph JOIN project_members pm ON pm.project_id=ph.project_id WHERE ph.id=? AND pm.user_id=?",(photo_id,user_id)): raise HTTPException(404,"Photo not found")
        c.execute("INSERT INTO photo_notes VALUES (?,?,?,?,?)", (note_id,photo_id,data.content,user_id,now()))
        return {"id": note_id}

@app.patch("/api/projects/{project_id}/general-note")
def update_general_note(project_id: str, data: GeneralNoteInput, request: Request):
    with db() as c:
        require_project(c,project_id,current_user_id(request)); c.execute("UPDATE projects SET general_note=?,updated_at=? WHERE id=?",(data.content,now(),project_id))
    return {"ok":True}

@app.get("/api/projects/{project_id}/albums/{album_id}/export")
def export_xlsx(project_id: str, album_id: str, request: Request):
    with db() as c:
        project=require_project(c,project_id,current_user_id(request))
        album=one(c,"SELECT name FROM albums WHERE id=? AND project_id=?",(album_id,project_id))
        if not album: raise HTTPException(404,"Album not found")
        rows=c.execute("""SELECT ph.file_name,ph.web_view_link,GROUP_CONCAT(n.content,CHAR(10)) notes FROM album_photos ap JOIN photos ph ON ph.id=ap.photo_id LEFT JOIN photo_notes n ON n.photo_id=ph.id WHERE ap.album_id=? GROUP BY ph.id ORDER BY ph.file_name""",(album_id,)).fetchall()
    workbook=Workbook(); selected=workbook.active; selected.title="Selected photos"; selected.append(["File name","Google Drive link","Notes"])
    for cell in selected[1]: cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="C86B67")
    for row in rows: selected.append([row["file_name"],row["web_view_link"],row["notes"] or ""])
    selected.column_dimensions["A"].width=24; selected.column_dimensions["B"].width=55; selected.column_dimensions["C"].width=58
    for row in selected.iter_rows():
        for cell in row: cell.alignment=Alignment(wrap_text=True,vertical="top")
    general=workbook.create_sheet("General notes"); general.append(["Project",project["name"]]); general.append(["Album",album["name"]]); general.append(["General edit notes",project["general_note"]]); general.column_dimensions["A"].width=22; general.column_dimensions["B"].width=100; general["B3"].alignment=Alignment(wrap_text=True,vertical="top")
    buffer=io.BytesIO(); workbook.save(buffer); buffer.seek(0)
    safe_filename = "wedding-photo-selection.xlsx"
    return StreamingResponse(buffer,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":f'attachment; filename="{safe_filename}"'})

init_db()
